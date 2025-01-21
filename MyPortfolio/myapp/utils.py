import smtplib
from email.message import EmailMessage
import string, re
import os, shutil
import datetime as dt
from bs4 import BeautifulSoup

import pandas as pd
from openpyxl import load_workbook
import bcrypt


#This function sends an email notification
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders

def send_mail(reciever, subject, message, sender, smtp="smtp.gmail.com", port=587, secret_key="cblibrpjtiqbohla",attachment=True,attachment_path=""):
    try:
        server = smtplib.SMTP(smtp, port)
        server.starttls()
        server.login(sender, secret_key)

        # Create the email message
        email = MIMEMultipart("alternative")
        email['From'] = sender
        email['To'] = reciever
        email['Subject'] = subject

        # Attach the HTML content to the email
        email.attach(MIMEText(message, 'html'))
        #If attachment is to be sent
        if attachment:
            with open(attachment, 'rb') as attachment: 
                part = MIMEBase('application', 'octet-stream') 
                part.set_payload(attachment.read()) 
                encoders.encode_base64(part) 
                part.add_header( 'Content-Disposition', f'attachment; filename={os.path.basename(csv_file_path)}', ) 
                email.attach(part)
        # Send the email
        server.send_message(email)
        server.quit()
        return "success"
    except Exception as e:
        print("Failed to send email notification: ", e)
        return "failed"

#To fetch and prepare the email template
# from bs4 import BeautifulSoup

def prepare_email_template_and_send(credentials_dict={}, automation_type="chat_app"):
    if 'sender' not in credentials_dict.keys():
        return False
    soup= """"""
    template_path = credentials_dict['invite_template_path']
    sender = credentials_dict['sender']
    password = credentials_dict['password']
    port = credentials_dict['port']
    host_domain = credentials_dict['host_domain_link']
    reciever = credentials_dict['reciever']
    username = credentials_dict['username']
    subject = credentials_dict['subject']
    attachment= False
    if automation_type=="chatapp":
    
        if credentials_dict['notification_type'] == "appointment_booking":
            subject =subject
            
        # Load the email template
        with open(template_path, 'r') as file:
            content = file.read()
            
        # Parse the HTML
        soup = BeautifulSoup(content, 'lxml')
        #Change title
        title=soup.find('title')
        title.string = "Email BOT by Prakhar"
        # Modify text in a specific tag
        header = soup.find('h3')
        header.string = f"This to notify you that your user id: {username} associated with {reciever}\n has been created!\n"
        header = soup.find('h4')
        header.string = "Welcome to the community!"
    elif automation_type == "portfolio":
        # Load the email template
        with open(template_path, 'r') as file:
            content = file.read()
            
        # Parse the HTML
        soup = BeautifulSoup(content, 'lxml')
        
        # Modify text in a specific tag
        message= credentials_dict['custom_message']
        signature = credentials_dict['signature']
        header = soup.find('h3')
        header.string = message.split("\n")[0]
        main_tag = soup.find('main')
        #header.string = signature
        for line in message.split("\n")[1:]:
            new_paragraph = soup.new_tag('p')
            new_paragraph.string = line
            main_tag.append(new_paragraph)
        #Add new content
        for line in signature.split("\n"):
            new_paragraph = soup.new_tag('p')
            new_paragraph.string = line
            main_tag.append(new_paragraph)
    elif automation_type =="log":
        attachment_filepath = credentials_dict['attachment_filepath']
        attachment= True
        with open(template_path, 'r') as file:
            content = file.read()
            
        # Parse the HTML
        soup = BeautifulSoup(content, 'lxml')
        
        # Modify text in a specific tag
        message= credentials_dict['custom_message']
        signature = credentials_dict['signature']
        header = soup.find('h3')
        header.string = message.split("\n")[0]
        main_tag = soup.find('main')
        #header.string = signature
        for line in message.split("\n")[1:]:
            new_paragraph = soup.new_tag('p')
            new_paragraph.string = line
            main_tag.append(new_paragraph)
        #Add new content
        for line in signature.split("\n"):
            new_paragraph = soup.new_tag('p')
            new_paragraph.string = line
            main_tag.append(new_paragraph)
        
    # Send email with HTML content
    result = send_mail(subject=subject, message=str(soup), sender=sender, secret_key=password, reciever=reciever, attachment= attachment, attachment_path=attachment_filepath)
    return result == "success"



#The function encrypts the string provided as an argument and returns it
def hash_password(password):
    # Generate a salt
    salt = bcrypt.gensalt()
    
    # Hash the password with the salt
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
    
    return hashed_password

# Example usage
# password = "anshu123"
# hashed_password = hash_password_bcrypt(password)
# print(f"Original password: {password}")
# print(f"Bcrypt hashed password: {hashed_password.decode('utf-8')}")


# This would be the hashed password stored in your database
hashed_password = b'$2b$12$K7Q16lOZ/.Jp6R.lXhQ..eFh5eYpQnVx8xNlp8AGvH2ZZZszAqC82'

# This is the password entered by the user
incoming_password = 'user_password'
def check_password(hashed_password, user_entered_password):
    # Check if the incoming password matches the hashed password
    if bcrypt.checkpw(user_entered_password.encode('utf-8'), hashed_password):
        print("Password matches!")
        return True
    else:
        print("Password does not match.")
        return False

#check_password(hashed_password=hashed_password, user_entered_password= incoming_password)

# This Function takes a list or a string as an input, and returns the lowercase values of the string/list elements after removing all the spaces from it.
# If the parameter  'remove_punc' is passed as True, the function will remove all the punctuations from the element/elements in addition to the aforementioned.
def format_headers(list_of_headers, remove_punc=False, custom_punctuations="", exclude_punc="", remove_all_spaces=True):
    try:
        new_headers = list_of_headers
        default_punctuations = string.punctuation
        if len(exclude_punc) != 0:
            default_punctuations = ''.join(char for char in default_punctuations if char not in exclude_punc)
        if len(custom_punctuations) != 0:
            default_punctuations = default_punctuations + custom_punctuations
        if type(list_of_headers) == list:
            if remove_punc:
                # if len(custom_punctuations) != 0:
                #     if len(exclude_punc) != 0:
                #         default_punctuations = [char for char in default_punctuations if char not in exclude_punc]
                new_headers = list(map(lambda x: str(x).translate
                (str.maketrans('', '', default_punctuations)), list_of_headers))


            elif custom_punctuations != "":
                new_headers = list(map(lambda x: str(x).translate
                (str.maketrans('', '', custom_punctuations)), list_of_headers))

            headers_to_format = list(map(lambda x: str(x).strip().lower(), new_headers))
            if remove_all_spaces:
                new_headers = list(map(lambda x: re.sub("\s+", "", x), headers_to_format))
            else:
                new_headers = headers_to_format
            return new_headers
        elif type(list_of_headers) == str:
            if remove_punc:


                new_headers = list_of_headers.translate(
                    str.maketrans('', '', default_punctuations + custom_punctuations))

            elif custom_punctuations:
                new_headers = list_of_headers.translate(str.maketrans('', '', default_punctuations))
            header_formatted = str(new_headers).strip().lower()
            if remove_all_spaces:
                new_headers = new_headers = re.sub("\s+", "", header_formatted)
            else:
                new_headers = header_formatted

            return new_headers
        else:
            print(
                "ERROR in format_headers function.\nThe input can only be of the type 'list'  or 'string'. The input datatype is : ",
                type(list_of_headers))
            return new_headers
    except Exception as ee:
        print("Error while formatting headers", ee)
        return list_of_headers


#print(format_headers("P* Rakhar", remove_punc=True, exclude_punc="", remove_all_spaces=False))


def remove_sheet_if_exists(excel_filepath, sheetname_to_be_removed="", create_if_not_exists=False):
    status = ""
    try:

        if not os.path.isfile(excel_filepath) or len(sheetname_to_be_removed) == 0:
            status = "failed"
            print("The Excel File Does Not Exist at the Path Provided!")
            return status
        wb = load_workbook(excel_filepath)
        existing_sheetnames = wb.sheetnames
        print("\nExisting Sheetnames : ", existing_sheetnames)
        if sheetname_to_be_removed in wb.sheetnames:

            wb.remove(wb[sheetname_to_be_removed])
            wb.save(excel_filepath)
            status = "success"
            print("Sheetname found in existing sheets,\nThe existing sheets after removal: ", wb.sheetnames)
            return status
        else:
            status = "success"
            print(f"The sheet by the name {sheetname_to_be_removed} does not exist! ")

    except Exception as exp:
        print(f"Utility function failed to delete the excel sheet!: {exp}")
        status = "failed"
        return status
    return status


# excel_filepath="C:/DCB SPARC/temperory/dummy output/all_reports_master.xlsx"
# print(remove_sheet_if_exists(excel_filepath,sheetname_to_be_removed='ALUPLRD FAILURE'))
# Thsi function takes input and output file or folder path and name(Full Path) as input and renames/moves the inputfile to the output location after renaming 
def rename_and_move(input_filepath, output_filename_and_path):
    status = ""
    try:
        print("Inside rename and move utility function!")
        if os.path.isfile(input_filepath) and not os.path.isfile(output_filename_and_path):

            # This code removes the file extention and returns the file path/name before that:
            # filename = "".join([word for word in (input_filepath.split(".")[:-1])])
            # filepath = re.sub("[\]","/",filename)
            # filename = re.split("[/]",filepath)[-1]
            # filepath= "".join((re.split("/",filepath))[:-1])
            # filepath = "".join(filepath)
            # print("filepath : ",filepath)
            # Rename the file 'old_file_name' to 'new_file_name'
            os.rename(input_filepath, output_filename_and_path)
            if os.path.isfile(output_filename_and_path):
                print("The File was renamed successfully to->" + output_filename_and_path)
                status = "success"
            else:
                status = "failed"
                print("The file could not be renamed!")
        elif os.path.isdir(input_filepath) and not os.path.isdir(output_filename_and_path):
            os.rename(input_filepath, output_filename_and_path)
            if os.path.isdir(output_filename_and_path):
                print("The Folder was renamed successfully to->" + output_filename_and_path)
                status = "success"
            else:
                status = "failed"
                print("The folder could not be renamed!")

        else:
            print(f"No File  or a Folder was found at the given path->{input_filepath}")
            status = "failed"
        return status
    except Exception as exp:
        print(f"Failed to rename the file/Folder!->{exp}")
        status = "failed"
        return status


# print("rename output : ",rename_and_move("C:/DCB SPARC/temperory/dummy output renamed",f"C:/DCB SPARC/temperory/dummy output"))#/repayment_IIFL_{dt.datetime.now().strftime("%m_%d_%Y__%H_%M")}_collection.txt"))
def create_dir_if_not_exists(folder_creation_path: str, foldername: str, overwrite=False):
    status = ""
    error_type = ""
    error_message = ""
    output = []
    new_folderpath = ""
    try:

        print("Inside Create dir Utility!")
        folder_creation_path = re.sub(r"\\", "/", folder_creation_path)
        print("folderpath : ", folder_creation_path)
        if os.path.isdir(folder_creation_path) and foldername:
            new_folderpath = os.path.join(folder_creation_path, foldername)
            new_folderpath = re.sub(r"\\", "/", new_folderpath)
            print("New Folderpath : ", new_folderpath)
            if os.path.isdir(new_folderpath):
                if overwrite:
                    shutil.rmtree(new_folderpath)
                    error_message = error_message + "->The Specified Directory Already Existed.It was Overwitten!"
                    os.mkdir(new_folderpath)
                    print("Created")
                else:
                    error_message = error_message + "->Folder Already Exists "
                    status = "success"

                    return status, "", [error_message, new_folderpath]

            else:
                os.mkdir(new_folderpath)
                print("Created")

        else:
            status = "failed"
            error_type = "invalid_input"
            error_message = error_message + "->The path to the directory where the new folder is to be created, does not exist !"
            return status, error_type, [error_message, ""]

        if os.path.isdir(new_folderpath):
            status = "success"

            return status, "", ["New Folder Created", new_folderpath]
        else:
            status = "failed"
            error_type = "functional_error"
            error_message = error_message + "->The New Folder Could Not Be Created!"
            return status, error_type, [error_message, ""]
    except Exception as e:
        status = "failed"
        error_type = "functional_error"
        error_message = error_message + f"Error in Create Dir utility function->The New Folder Could Not Be Created! ->{e}"
        return status, error_type, [error_message, ""]


def search_files(input_filepath=""):
    status = ""
    error_type = ""
    error_message = ""
    output = []
    try:
        print("Inside Search files utility function")
        if len(input_filepath):

            input_filepath = re.sub(r"\\", "/", input_filepath)
            print("Renamed path : ", input_filepath)
        else:
            status = "failed"
            error = "file_not_found"
            error_message = "Input Filepath Can Not Be Empty!"
            return status, error_type, [error_message]
        if os.path.isfile(input_filepath):
            status = "success"
            error_type = ""
            return status, error_type, [error_message]
        elif os.path.isdir(input_filepath):
            status = "success"

            return status, error_type, [error_message]
        else:
            status = "failed"
            error_message = error_message + f"The file was not present at the provided path->{input_filepath}"
            return status, error_type, [error_message]

    except Exception as e:
        status = "failed"
        error_type = error_message + f"Error in Search Files Utility Function!File_not_found->{e}"
        return status, error_type, output


# print("search : ",search_files(input_filepath="C:\DCB SPARC\BAAR Studio SPARC\Readme.txt"))
# print("After Retrying : ",rt.retry_until_condition_is_satisfied(function_name=search_files,argument_list=["C:\DCB SPARC\\temperory\dummy input\collection.txt"],no_of_retries=4,time_until_retry=20))

# print("Folder creation status : ",create_dir_if_not_exists(folder_creation_path="C:/DCB SPARC/temperory/dummy input",foldername="Newly Created"))

def check_repayment_file_format(input_filepath, file_format="CLPCODE_yyyy_mm_dd_collection_lotno"):
    status = ""
    error_type = ""
    error_message = ""
    output = []

    # Prescribed file format attributes
    file_format_attributes_list = re.split("_", file_format.lower())
    # filename with extension
    input_filename = re.sub(r"\\", "/", input_filepath.strip().lower())
    input_filename = (re.split(r"\/", input_filename))[-1]
    print("\nInput filename with extention :", input_filename)
    # Just filename without extension
    input_filename_raw = input_filename.split(".")[0]
    print("\nInput filename raw :", input_filename_raw)
    # Just File extension
    input_file_extension = input_filename.split(".")[-1]
    input_filename_attributes_list = re.split("_", input_filename_raw)
    print("Prescribed Format attributes : ", file_format_attributes_list)
    print("Input File attributes list : ", input_filename_attributes_list)
    # Performing Validations

    # If the file names have same number of attributes
    if len(file_format_attributes_list) != len(input_filename_attributes_list):
        return "failed", "invalid_name", [
            "The number of attributes provided in the collection filename do not match with the prescribed attributes",
            output]
    # If the file extension is correct
    if input_file_extension != "txt":
        return "failed", "input_file_format", ["Only .txt files are supported for this process", output]
    # extracting parameters from filename
    if "collection" in input_filename_raw:
        # Checking via regex
        file_format_checker_regex = "[A-Za-z]+[_]{1}(\d){4}[_]{1}(\d){2}[_]{1}(\d){2}[_]{1}collection[_]{1}[A_Za-z]{1}(\d){1}[\.]{1}txt"
        check_format = re.fullmatch(file_format_checker_regex, input_filename)
        if check_format is None:
            return "failed", "invalid_name", [
                "The number of attributes provided in the collection filename do not match with the prescribed attributes",
                output]
        else:
            filename_verified = check_format.group()
            # Checking if the file date is equal to current date
            date_raw = "-".join(input_filename_attributes_list[1:4])
            date_today = dt.datetime.today().strftime("%Y-%m-%d")
            print("\n Date Today :", date_today)
            print("Date Raw : ", date_raw)
            if date_raw != date_today:
                error_type = "date_error"
                error_message = "The date in the provided file must be the current date!"
                return "failed", error_type, [error_message, output]
            if filename_verified == input_filename:

                # append the filename
                output.append(input_filename)
                # Append CLP Code
                output.append(input_filename_attributes_list[0])
                # append_date
                output.append(date_raw)
                # Append filetype
                output.append(input_filename_attributes_list[4])
                # append lot no
                output.append(input_filename_raw.split("_")[-1])
                # Returning the outputs
                return "success", "", ["all validations done", output]
            else:
                return "failed", "functional_error", ["Please check the 'check file format' utility function!"]

    else:
        return "failed", "invalid_name", [
            "The attribute provided in the collection filename do not match with the prescribed attributes->The name should mention 'collection' as an identifier",
            output]


# print("File Format Validation status :",check_repayment_file_format("C:\DCB SPARC\temperory\dummy input\IIFL_2024_05_05_collection_l1.txt"))


def write_text_file(input_filepath, data, mode="w+", success_message="", failure_message="",
                    separator="\nprocess_tracker :-> \n"):
    status = ""
    try:
        # Writing the input in a text file
        with open(input_filepath, mode) as input:
            input.write(separator + data)
            # input.write("\n process_tracker : \n"+process_tracker_stringify)
            input.close()
            print(success_message)
            return "success"
    except Exception as ee:
        print(f"{failure_message}->{ee}")
        return "failed"


def get_processed_account_numbers(path_to_master_excel_folder: str, processed_lot_nos="1,2", sheetname="Account",
                                  column_name="account_numbers"):
    account_numbers_master_list = []
    # Checking if the folder for the current day master report exists
    processed_lot_no_list = processed_lot_nos.split(",")
    today_foldername = dt.datetime.now().strftime('%d_%b_%Y')
    print("Inside Processed Account Numbers", processed_lot_no_list)
    folder_creation_path = re.sub(r"\\", "/", path_to_master_excel_folder)
    print("folderpath : ", path_to_master_excel_folder)

    if os.path.isdir(path_to_master_excel_folder) and len(os.listdir(path_to_master_excel_folder)) != 0:
        all_master_folders = os.listdir(path_to_master_excel_folder)
        print("folderpath\n", path_to_master_excel_folder, "\noutput\n ", all_master_folders)
    else:
        return "", "", [
            f"The Folder Containing Master Reports for {dt.datetime.now().strftime('%d_%b_%Y')} Has Not Been Created Yet!",
            []]
    for foldername in all_master_folders:
        if len(foldername.split("_")) < 5:
            continue
        lot_no_raw = foldername.split("_")[1]
        lot_no = str(int(lot_no_raw[-1]))
        print("here : ", foldername)
        if (lot_no not in processed_lot_no_list) and (today_foldername in foldername) and ("failed" not in foldername):
            print("here 3: ", foldername)
            list_of_files = os.listdir(os.path.join(path_to_master_excel_folder, foldername))
            for file in list_of_files:
                full_filepath = os.path.join(path_to_master_excel_folder, foldername, file)
                # print("here2 : ", full_filepath)

                account_numbers_df = pd.read_excel(full_filepath, sheet_name=sheetname)
                # print("acc nos\n : ",account_numbers_df)
                account_numbers_list = list(account_numbers_df[column_name].unique())
                account_numbers_master_list.extend(list(set(account_numbers_list)))
        print("Total Number Of Accounts Processed Today : ", len(account_numbers_master_list))

    return "success", "", [
        f"existing_account_numbers_fetched", account_numbers_master_list]
    # print("\nFinal :", account_numbers_master_list)

# print("Final : \n",get_processed_account_numbers("C:/DCB SPARC/collection process automation/IIFL/all_reports_master"))
